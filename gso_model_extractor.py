"""
gso_model_extractor.py

Reads GSO (Gulf Conformity Mark / gso.org.sa) product page URLs from an Excel
file, scrapes each page for its "Model Numbers" table AND its "Product
Specifications" fields, and writes everything back into the workbook.

WHY MODELS WEREN'T BEING PULLED UP -- FIXES IN THIS VERSION
---------------------------------------------------------------
Three real bugs were found by tracing the code (not by running it live --
see the note below on that):

  1. STRICT HEADER MATCH: find_model_table() only matched a table whose
     first header cell was the EXACT text "number" (case-folded). Tested
     against plausible real variants -- "Model Number", "Product Number",
     "No." -- and every one of them silently returned zero models instead
     of erroring. This alone could fully explain "doesn't pull up the
     model no" if the real site's header text is anything other than the
     literal word "Number". FIX: match now checks whether the header cell
     contains any of MODEL_HEADER_KEYWORDS (default: "number", "model"),
     case-insensitively, instead of requiring an exact match.

  2. PREMATURE "STABLE" EXIT AT ZERO: the scroll/retry loop's stability
     check couldn't tell "stable because we're genuinely done" apart from
     "stable because nothing has rendered yet". Simulated it directly: a
     sequence where real data only appears on scroll round 6 gets cut off
     after round 3 regardless -- the loop was set up to handle exactly
     this lazy-load case but was giving up on it in under 2 seconds. FIX:
     the stable-count early exit is now only trusted once count > 0.
     While count is still 0, the loop keeps scrolling through every
     round available (up to MAX_SCROLL_ROUNDS) instead of accepting an
     early "stable" zero as final.

  3. EARLY RETURN ON THE INITIAL TABLE WAIT: if wait_for_selector("table")
     timed out (10s), the original code returned immediately, never even
     entering the scroll/retry loop that was built to handle slow-loading
     tables. FIX: that failure is now just logged and falls through into
     the same loop as everything else -- a page with no table yet reads
     as count=0, which fix #2 above now handles patiently instead of
     bailing out.

NEW IN THIS VERSION: PRODUCT SPECIFICATION FIELDS
---------------------------------------------------
Added extraction for: Product Category, Product Scope, Product HS Code,
Product Description, Trademark/Brand, Manufacturer, and Manufacturing
Country (the site's own label -- you asked for "manufacturer country";
the output column is named "Manufacturer Country" but pulls from the
site's "Manufacturing Country" field, since that's what's actually on
the page).

Unlike the earlier "I couldn't verify any of this" version, this part
WAS checked against a live page (product 84338,
https://www.gso.org.sa/nb/p/84338). That fetch confirmed:
  - A "Product Specifications" section exists with exactly these fields,
    each appearing as a label immediately followed by its value (e.g. the
    text "Product Category" is immediately followed by the text
    "Household air conditioners"), same as the Model Numbers section --
    consistent with this whole product page being built from div/grid
    blocks rather than real <table> markup.
  - The section ends where "Countries Considered" (then "Model Numbers")
    begins, so that's used as the stop boundary.
  - Field extraction uses a structural attempt first (checks for a real
    <table> row or <dt>/<dd> pair with a matching label, in case some
    other product page IS built with real markup), and falls back to the
    line-based label->next-line text scan when nothing structural
    matches -- same two-stage pattern already used for Model Numbers.

This does NOT guarantee every product page on the site is laid out
identically -- only that this pattern is confirmed for at least one real
page. The DEBUG_ROWS diagnostics below now also fire if all spec fields
come back empty, so any page that doesn't match this layout will be
visible in the console instead of silently returning blanks.

WHAT I STILL COULDN'T VERIFY
------------------------------
I don't have your actual Excel file, so column names / row counts are
untested. And BASE_URL = "https://www.gso.org.sa/nb/p/{}" is confirmed
correct as a pattern (84338 resolved directly to a real product page with
no redirect), but I only checked it for that one product number.

NEW: DIAGNOSTICS
------------------
When a row ends up with zero models AND/OR all spec fields blank after
the full scroll/retry loop, and DEBUG_ROWS hasn't been exhausted yet,
this version prints:
  - the page's final resolved URL (catches silent redirects, e.g. to a
    login or error page)
  - the page title
  - EVERY <table> found on the page, with its first row's cell text
  - the raw HTML around any "Model Numbers" text node
This is capped at DEBUG_ROWS rows (default 5) so a full run doesn't flood
the console.

Set TEST_ROWS to a small number (e.g. 3) to run against just the first
few rows while you're checking whether these fixes actually resolve it
for your sheet, before committing to a full run.

USAGE
-----
1. Install dependencies (one time, in the SAME Python environment VS Code
   is using -- check the interpreter shown in the bottom-right corner):

       pip install playwright beautifulsoup4 openpyxl
       python -m playwright install chromium

2. Either edit INPUT_FILE below, OR run with a filename argument:

       python gso_model_extractor.py GSO_test_small.xlsx

3. Output is saved as "<original_name>_filled.xlsx" next to the input file.

If you still see missing rows after this, set HEADLESS = False and re-run
on just the problem row(s) so you can watch what the page actually does --
that'll tell us whether it's a different pagination pattern we need to
target specifically.
"""

import re
import sys
import time
from pathlib import Path

from bs4 import BeautifulSoup
import openpyxl

try:
    from playwright.sync_api import sync_playwright
except ImportError:
    sys.exit(
        "Playwright is not installed. Run:\n"
        "    pip install playwright beautifulsoup4 openpyxl\n"
        "    python -m playwright install chromium"
    )

# ---------------------------------------------------------------------------
# CONFIG
# ---------------------------------------------------------------------------

INPUT_FILE = "GSO_Batch2_Input.xlsx"
SHEET_NAME = None

URL_COLUMN = "GSO Product Page"
PRODUCT_NUMBER_COLUMN = "Product GCTS"

OUTPUT_MODEL_COLUMN = "Model No. (extracted)"
OUTPUT_STATUS_COLUMN = "Status"

BASE_URL = "https://www.gso.org.sa/nb/p/{}"
BASE_DELAY = 1.0
PAGE_TIMEOUT_MS = 30000
MAX_RETRIES = 2

MAX_SCROLL_ROUNDS = 20        # safety cap on scroll/click iterations
SCROLL_WAIT_MS = 500          # wait after each scroll before re-checking
STABLE_ROUNDS_REQUIRED = 2    # consecutive unchanged counts before stopping (only once count > 0 -- see fix #2)

# The first header cell of the target table must contain one of these
# (case-insensitive) to be treated as the model-numbers table. Broadened
# from an exact match on "number" -- see fix #1. Add to this list if the
# diagnostics below reveal a different real header, e.g. "no." or "model no".
MODEL_HEADER_KEYWORDS = ["number", "model"]

# Product Specifications fields to pull, mapped to the output column header
# you want in the spreadsheet. Keys are the EXACT label text confirmed on
# the live page (product 84338); values are your requested output header.
# "Manufacturing Country" is the site's real label -- you asked for
# "manufacturer country", so that's what the output column is named, but
# it's matched against the site's actual "Manufacturing Country" text.
OUTPUT_SPEC_COLUMNS = {
    "Product Category": "Product Category",
    "Product Scope": "Product Scope",
    "Product HS Code": "Product HS Code",
    "Product Description": "Product Description",
    "Trademark/Brand": "Trademark/Brand",
    "Manufacturer": "Manufacturer",
    "Manufacturing Country": "Manufacturer Country",
}
PRODUCT_SPEC_FIELDS = list(OUTPUT_SPEC_COLUMNS.keys())

# Text markers that end the Product Specifications block on the live page.
SPEC_STOP_MARKERS = {"countries considered", "model numbers", "technical specifications"}

ONE_ROW_PER_MODEL = False
HEADLESS = True

# --- Debugging / diagnostics ---
DEBUG_ROWS = 5     # print full diagnostics (URL, title, all table headers) for
                   # the first N rows that come back with zero models or all
                   # spec fields empty. Set to 0 to disable, or a big number
                   # to always see it.
TEST_ROWS = None   # set to a small int (e.g. 3) to only process the first N
                   # rows this run, for quick iteration while debugging.

# --- Web-app overrides (used by app.py; harmless for normal CLI use) ---
import os as _os
if _os.environ.get("WEBAPP_TEST_ROWS"):
    TEST_ROWS = int(_os.environ["WEBAPP_TEST_ROWS"])
if _os.environ.get("WEBAPP_INPUT_FILE"):
    INPUT_FILE = _os.environ["WEBAPP_INPUT_FILE"]

# ---------------------------------------------------------------------------

_debug_rows_shown = 0  # module-level counter so DEBUG_ROWS caps output across the whole run


def find_model_table(soup):
    for table in soup.find_all("table"):
        rows = table.find_all("tr")
        if not rows:
            continue
        first_row_cells = rows[0].find_all(["td", "th"])
        if not first_row_cells:
            continue
        first_cell_text = first_row_cells[0].get_text(strip=True).lower()
        if any(keyword in first_cell_text for keyword in MODEL_HEADER_KEYWORDS):
            return table
    return None


def extract_models_from_html(html):
    """
    Two-stage extraction:
      1. Structural: look for a real <table> whose header matches
         MODEL_HEADER_KEYWORDS (works if the site DOES use a real table).
      2. Fallback (heuristic): if no table matched, fetching
         https://www.gso.org.sa/nb/p/84888 directly showed real model data
         ('EBE4302C-B') present under a 'Model Numbers' section with NO
         scrolling needed at all -- strong evidence that section isn't a
         <table> element on this site, but a div/grid-based layout instead
         (consistent with the site's custom 'gso-uicore' component
         framework). So as a fallback, we locate the 'Model Numbers'
         section by its heading TEXT instead of by tag structure, take
         everything up to the next section heading, and strip out the
         known column-label words ('Number', 'Notes', 'Barcode').

    KNOWN LIMITATION of the fallback: if a product's Notes or Barcode
    column is actually populated (not blank), those values get pulled in
    too, since plain text can't distinguish which column a value belongs
    to without knowing the real markup. Results from this path are
    tagged 'heuristic' (see caller) precisely so they can be spot-checked
    rather than trusted blindly -- this needs the debug_dump_tables /
    debug_dump_model_section output confirmed against the real DOM to
    tighten up further.

    Returns (models, source) where source is "table", "heuristic", or "none".
    """
    soup = BeautifulSoup(html, "html.parser")
    table = find_model_table(soup)
    models = []
    if table is not None:
        for row in table.find_all("tr")[1:]:
            cells = row.find_all(["td", "th"])
            if not cells:
                continue
            val = cells[0].get_text(strip=True)
            if val and not any(keyword in val.lower() for keyword in MODEL_HEADER_KEYWORDS) and val not in models:
                models.append(val)
    if models:
        return models, "table"

    heuristic_models = extract_models_from_html_fallback(html)
    if heuristic_models:
        return heuristic_models, "heuristic"

    return [], "none"


def extract_models_from_html_fallback(html):
    """
    Text-based fallback for when the Model Numbers section isn't a real
    <table> (see extract_models_from_html docstring for why this exists).
    Finds the 'Model Numbers' section by its heading text, grabs
    everything up to the next known section heading, and returns whatever
    isn't one of the known column-label words.
    """
    soup = BeautifulSoup(html, "html.parser")
    full_text = soup.get_text("\n", strip=True)
    lines = [l.strip() for l in full_text.split("\n") if l.strip()]

    # The real page showed 'Model Numbers' twice in a row (once as the page
    # section heading, once as the card/block's own title) -- prefer the
    # SECOND occurrence as the actual data block start, but fall back to
    # the first if it only appears once.
    start_idx = None
    count_seen = 0
    for i, line in enumerate(lines):
        if line.lower() == "model numbers":
            count_seen += 1
            if count_seen == 2:
                start_idx = i
                break
    if start_idx is None:
        for i, line in enumerate(lines):
            if line.lower() == "model numbers":
                start_idx = i
                break
    if start_idx is None:
        return []

    # Stop at the next section -- confirmed real heading text from the live
    # page. Also capped at 40 lines out as a safety net in case none of
    # these markers appear (e.g. a differently-structured product page).
    STOP_MARKERS = {"technical specifications", "additional technical specifications", "product photos"}
    end_idx = len(lines)
    for i in range(start_idx + 1, min(start_idx + 40, len(lines))):
        if lines[i].lower() in STOP_MARKERS:
            end_idx = i
            break

    block = lines[start_idx:end_idx]
    known_labels = {"model numbers"} | {k.lower() for k in MODEL_HEADER_KEYWORDS} | {"notes", "barcode"}
    candidates = [l for l in block if l.lower() not in known_labels]

    seen = set()
    result = []
    for c in candidates:
        if c not in seen:
            seen.add(c)
            result.append(c)
    return result


def extract_specs_from_structure(soup):
    """
    Structural attempt at pulling Product Specification fields: checks for
    a real <table> row (label cell + value cell) or a <dt>/<dd> pair whose
    label matches one of PRODUCT_SPEC_FIELDS. This is tried first in case
    some product pages (unlike the one checked live) use real markup for
    this section. Returns a dict of whatever it found (may be partial or
    empty) -- extract_product_specs() below fills in the rest via the
    text-based fallback, which IS confirmed to match the live page layout.
    """
    result = {}

    for table in soup.find_all("table"):
        for row in table.find_all("tr"):
            cells = row.find_all(["td", "th"])
            if len(cells) < 2:
                continue
            label = cells[0].get_text(strip=True)
            value = cells[1].get_text(strip=True)
            for field in PRODUCT_SPEC_FIELDS:
                if label.lower() == field.lower() and value:
                    result[field] = value

    for dl in soup.find_all("dl"):
        dts = dl.find_all("dt")
        dds = dl.find_all("dd")
        for dt, dd in zip(dts, dds):
            label = dt.get_text(strip=True)
            value = dd.get_text(strip=True)
            for field in PRODUCT_SPEC_FIELDS:
                if label.lower() == field.lower() and value:
                    result[field] = value

    return result


def extract_specs_from_text(lines):
    """
    Text-based extraction for Product Specification fields, confirmed
    against the live page for product 84338: the section starts at the
    'Product Specifications' heading, and each field name is immediately
    followed (as the next non-empty line) by its value, e.g.:

        Product Category
        Household air conditioners

    Stops at SPEC_STOP_MARKERS (Countries Considered / Model Numbers /
    Technical Specifications), matching what the live page showed right
    after the specification fields end.
    """
    result = {field: "" for field in PRODUCT_SPEC_FIELDS}

    start_idx = None
    for i, line in enumerate(lines):
        if line.lower() == "product specifications":
            start_idx = i
            break
    if start_idx is None:
        return result

    n = len(lines)
    i = start_idx + 1
    while i < n:
        low = lines[i].lower()
        if low in SPEC_STOP_MARKERS:
            break
        for field in PRODUCT_SPEC_FIELDS:
            if low == field.lower():
                j = i + 1
                while j < n and not lines[j].strip():
                    j += 1
                if j < n:
                    result[field] = lines[j].strip()
                break
        i += 1

    return result


def extract_product_specs(html):
    """
    Combined extraction for all PRODUCT_SPEC_FIELDS: tries the structural
    approach first, then fills any still-blank fields from the text-based
    approach (which is the one confirmed against the live page). Returns a
    dict keyed by the site's field names (see OUTPUT_SPEC_COLUMNS for the
    output column mapping).
    """
    soup = BeautifulSoup(html, "html.parser")
    result = {field: "" for field in PRODUCT_SPEC_FIELDS}

    structural = extract_specs_from_structure(soup)
    for field, value in structural.items():
        if value:
            result[field] = value

    missing = [f for f in PRODUCT_SPEC_FIELDS if not result.get(f)]
    if missing:
        full_text = soup.get_text("\n", strip=True)
        lines = [l.strip() for l in full_text.split("\n") if l.strip()]
        text_based = extract_specs_from_text(lines)
        for f in missing:
            if text_based.get(f):
                result[f] = text_based[f]

    return result


def debug_dump_tables(html):
    """
    Diagnostic: list every <table> on the page and its first row's cell
    text, so a real header-text mismatch or a wrong/error page is visible
    directly in the console instead of guessed at.
    """
    soup = BeautifulSoup(html, "html.parser")
    tables = soup.find_all("table")
    if not tables:
        print("      [debug] No <table> elements found on this page at all "
              "-- consistent with the Model Numbers / spec sections being a "
              "div/grid layout rather than real tables (see debug_dump_model_section below).")
        return
    print(f"      [debug] {len(tables)} <table> element(s) on this page "
          f"(first row's cell text for each, up to 6 columns):")
    for i, table in enumerate(tables, start=1):
        rows = table.find_all("tr")
        if not rows:
            print(f"      [debug]   table #{i}: (no rows)")
            continue
        first_row_cells = rows[0].find_all(["td", "th"])
        cell_texts = [c.get_text(strip=True) for c in first_row_cells[:6]]
        print(f"      [debug]   table #{i}: {cell_texts}")


def debug_dump_model_section(html):
    """
    Diagnostic: find any element whose OWN text is exactly 'Model Numbers'
    (or contains 'Barcode'/case-insensitive 'Number' as a plausible nearby
    label), and print the RAW HTML of a reasonably-sized ancestor around
    it. Unlike debug_dump_tables (which only looks at <table> tags), this
    finds the section regardless of what element type it's actually built
    from -- divs, lists, custom components, whatever.
    """
    soup = BeautifulSoup(html, "html.parser")
    targets = soup.find_all(string=lambda s: s and s.strip().lower() == "model numbers")
    if not targets:
        print("      [debug] No element with the exact text 'Model Numbers' found on this page.")
        return
    print(f"      [debug] {len(targets)} element(s) with text 'Model Numbers' found. "
          f"Raw HTML around each (walking up to a reasonably-sized container):")
    for i, text_node in enumerate(targets, start=1):
        el = text_node.parent
        container = el
        for _ in range(3):
            if container.parent is not None:
                container = container.parent
        snippet = str(container)
        if len(snippet) > 1500:
            snippet = snippet[:1500] + " ...[truncated]"
        print(f"      [debug]   --- occurrence #{i} (tag=<{el.name}>, class={el.get('class')}) ---")
        print(f"      [debug]   {snippet}")


def extract_status(html):
    soup = BeautifulSoup(html, "html.parser")
    page_text = soup.get_text(" ", strip=True)

    if re.search(r"invalid product number", page_text, re.I):
        return "Invalid product number"

    m = re.search(r"Activation will expire on ([\d-]+)", page_text)
    if m:
        return f"Active (expires {m.group(1)})"
    if re.search(r"does not have an active conformity certificate", page_text, re.I) \
            or re.search(r"no active certificate", page_text, re.I):
        return "No active certificate"
    return "Unknown"


def click_if_present(page, texts):
    """Try clicking any button/link matching one of the given texts. Returns True if clicked."""
    for text in texts:
        try:
            locator = page.locator(f"button:has-text('{text}'), a:has-text('{text}')").first
            if locator.count() > 0 and locator.is_visible():
                locator.click(timeout=2000)
                return True
        except Exception:
            continue
    return False


def load_all_model_rows(page):
    """
    Scroll and click through any lazy-load/pagination controls until the
    model table's row count stops growing. Returns the final HTML.

    FIX #3: the original code returned immediately if the initial
    wait_for_selector("table") timed out, skipping the retry loop below
    entirely. Now a timeout here is just logged and falls through --
    the loop's first iteration reads the page (however loaded so far) and
    proceeds from there.

    FIX #2: the stable-count early exit is now only trusted once we've
    actually seen count > 0. A stable count of exactly 0 no longer counts
    as "done" -- it keeps scrolling through every round up to
    MAX_SCROLL_ROUNDS, since a false/premature zero was the likely cause
    of rows silently coming back empty.
    """
    try:
        page.wait_for_selector("table", timeout=10000)
    except Exception:
        pass  # don't bail -- fall through to the loop below and keep trying

    prev_count = -1
    stable = 0

    for _ in range(MAX_SCROLL_ROUNDS):
        html = page.content()
        count = len(extract_models_from_html(html)[0])

        if count > 0:
            if count == prev_count:
                stable += 1
                if stable >= STABLE_ROUNDS_REQUIRED:
                    break
            else:
                stable = 0
        prev_count = count

        page.mouse.wheel(0, 3000)
        page.evaluate("window.scrollTo(0, document.body.scrollHeight)")
        click_if_present(page, ["Show more", "Load more", "Next", "المزيد", "التالي"])
        page.wait_for_timeout(SCROLL_WAIT_MS)

    return page.content()


def fetch_and_extract(page, url):
    """
    Load a URL, exhaust lazy-loading, and return (html, models, status,
    source, specs) where specs is a dict keyed by PRODUCT_SPEC_FIELDS.
    """
    global _debug_rows_shown

    for attempt in range(1, MAX_RETRIES + 1):
        try:
            page.goto(url, timeout=PAGE_TIMEOUT_MS, wait_until="networkidle")
            page.wait_for_timeout(500)
            html = load_all_model_rows(page)
            models, source = extract_models_from_html(html)
            status = extract_status(html)
            specs = extract_product_specs(html)

            specs_all_blank = not any(specs.values())
            if _debug_rows_shown < DEBUG_ROWS and (source != "table" or specs_all_blank):
                _debug_rows_shown += 1
                try:
                    print(f"      [debug] resolved URL: {page.url}")
                    print(f"      [debug] page title:   {page.title()}")
                except Exception:
                    pass
                if specs_all_blank:
                    print("      [debug] all Product Specification fields came back empty")
                debug_dump_tables(html)
                debug_dump_model_section(html)

            if status == "Invalid product number":
                return html, [], status, "none", {f: "" for f in PRODUCT_SPEC_FIELDS}
            return html, models, status, source, specs
        except Exception as e:
            print(f"  [warn] {url} -> {e} (attempt {attempt})")
            time.sleep(2 * attempt)
    return None, [], "FETCH FAILED", "none", {f: "" for f in PRODUCT_SPEC_FIELDS}


def resolve_columns(ws):
    headers = {}
    for col in range(1, ws.max_column + 1):
        val = ws.cell(row=1, column=col).value
        if val:
            headers[str(val).strip()] = col
    return headers


def main():
    input_file = sys.argv[1] if len(sys.argv) > 1 else INPUT_FILE
    in_path = Path(input_file)
    if not in_path.exists():
        sys.exit(f"Input file not found: {in_path.resolve()}")

    print(f"Input file : {in_path.resolve()}")

    wb = openpyxl.load_workbook(in_path)
    ws = wb[SHEET_NAME] if SHEET_NAME else wb.active
    print(f"Sheet      : {ws.title}")

    headers = resolve_columns(ws)
    print(f"Columns    : {list(headers)}")

    url_col = headers.get(URL_COLUMN)
    num_col = headers.get(PRODUCT_NUMBER_COLUMN)

    if url_col is None and num_col is None:
        sys.exit(
            f"\nERROR: could not find column '{URL_COLUMN}' or "
            f"'{PRODUCT_NUMBER_COLUMN}' in row 1.\nFound columns: {list(headers)}"
        )

    model_col = headers.get(OUTPUT_MODEL_COLUMN)
    if model_col is None:
        model_col = ws.max_column + 1
        ws.cell(row=1, column=model_col, value=OUTPUT_MODEL_COLUMN)

    status_col = headers.get(OUTPUT_STATUS_COLUMN)
    if status_col is None:
        status_col = ws.max_column + 1
        ws.cell(row=1, column=status_col, value=OUTPUT_STATUS_COLUMN)

    # Resolve/create one output column per Product Specification field.
    spec_cols = {}
    for site_field, out_header in OUTPUT_SPEC_COLUMNS.items():
        col = headers.get(out_header)
        if col is None:
            col = ws.max_column + 1
            ws.cell(row=1, column=col, value=out_header)
            headers[out_header] = col
        spec_cols[site_field] = col

    work_rows = []
    for row in range(2, ws.max_row + 1):
        url = None
        if url_col:
            v = ws.cell(row=row, column=url_col).value
            if v:
                url = str(v).strip()
        if url is None and num_col:
            v = ws.cell(row=row, column=num_col).value
            if v:
                url = BASE_URL.format(str(v).strip())
        if url:
            work_rows.append((row, url))

    print(f"Rows found : {len(work_rows)} (of {ws.max_row - 1} data rows in sheet)")

    if TEST_ROWS is not None:
        work_rows = work_rows[:TEST_ROWS]
        print(f"TEST_ROWS is set to {TEST_ROWS} -- only processing that many rows this run.")

    print()

    if not work_rows:
        sys.exit("No URLs/product numbers found — check your column settings.")

    ok, invalid, failed = 0, 0, 0

    with sync_playwright() as p:
        browser = p.chromium.launch(headless=HEADLESS)
        context = browser.new_context(
            user_agent=(
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
                "(KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36"
            )
        )
        page = context.new_page()

        for i, (row, url) in enumerate(work_rows, start=1):
            print(f"[{i}/{len(work_rows)}] {url}")
            html, models, status, source, specs = fetch_and_extract(page, url)

            if html is None:
                ws.cell(row=row, column=status_col, value="FETCH FAILED")
                failed += 1
                time.sleep(BASE_DELAY)
                continue

            status_out = status if source != "heuristic" else f"{status} [heuristic]"
            ws.cell(row=row, column=status_col, value=status_out)
            ws.cell(row=row, column=model_col, value=", ".join(models))

            for site_field, col in spec_cols.items():
                ws.cell(row=row, column=col, value=specs.get(site_field, ""))

            if status == "Invalid product number":
                invalid += 1
            elif models:
                ok += 1
                print(f"  -> {len(models)} model(s) found" + (" [heuristic -- spot-check]" if source == "heuristic" else ""))
            else:
                print(f"  [note] no model numbers found (status: {status})")

            filled_specs = sum(1 for v in specs.values() if v)
            print(f"  -> {filled_specs}/{len(PRODUCT_SPEC_FIELDS)} spec field(s) filled")

            time.sleep(BASE_DELAY)

        browser.close()

    out_path = in_path.with_name(in_path.stem + "_filled.xlsx")

    if ONE_ROW_PER_MODEL:
        wb = _explode_rows(wb, ws.title, model_col)

    wb.save(out_path)

    print(
        f"\nDone. {ok} rows with models, {invalid} invalid, "
        f"{failed} fetch failures out of {len(work_rows)} total."
    )
    print(f"Saved: {out_path.resolve()}")


def _explode_rows(wb, sheet_name, model_col):
    ws = wb[sheet_name]
    new_wb = openpyxl.Workbook()
    new_ws = new_wb.active
    new_ws.title = ws.title

    for col in range(1, ws.max_column + 1):
        new_ws.cell(row=1, column=col, value=ws.cell(row=1, column=col).value)

    out_row = 2
    for row in range(2, ws.max_row + 1):
        models_cell = ws.cell(row=row, column=model_col).value or ""
        models = [m.strip() for m in models_cell.split(",") if m.strip()]
        if not models:
            models = [""]
        for model in models:
            for col in range(1, ws.max_column + 1):
                val = ws.cell(row=row, column=col).value
                if col == model_col:
                    val = model
                new_ws.cell(row=out_row, column=col, value=val)
            out_row += 1

    return new_wb


if __name__ == "__main__":
    main()
